from fastapi import FastAPI, Request
from fastapi.responses import StreamingResponse
from fastapi.middleware.cors import CORSMiddleware
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import json
import io

app = FastAPI()

# Cho phép Power BI gọi API
app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_methods=["*"],
    allow_headers=["*"],
)

# ─── Màu sắc giống Gantt HTML ───────────────────────────────
COLOR = {
    "header_bg":   "1F4E79",
    "header_fg":   "FFFFFF",
    "total_bg":    "DDEBF7",
    "total_fg":    "1F4E79",
    "hc_bg":       "BDD7EE",
    "hc_fg":       "0033CC",
    "ot_bg":       "F8CBAD",
    "ot_fg":       "C00000",
    "date_bg":     "9CC3E6",
    "done_bg":     "C6EFCE",
    "done_fg":     "006100",
    "todo_bg":     "FFC7CE",
    "todo_fg":     "9C0006",
    "inprog_bg":   "D9E1F2",
    "inprog_fg":   "305496",
    "footer_bg":   "1F4E79",
    "footer_fg":   "FFFFFF",
    "footer_date_bg": "4472C4",
}

def make_fill(hex_color):
    return PatternFill(start_color=hex_color, end_color=hex_color, fill_type="solid")

def make_font(hex_color, bold=False, size=11):
    return Font(color=hex_color, bold=bold, size=size, name="Segoe UI")

def make_border():
    side = Side(style="thin", color="CCCCCC")
    return Border(left=side, right=side, top=side, bottom=side)

def make_align(horizontal="left"):
    return Alignment(horizontal=horizontal, vertical="center", wrap_text=False)


@app.post("/export-gantt")
async def export_gantt(request: Request):
    """
    Nhận JSON từ Power BI HTML visual, tạo file .xlsx có màu và trả về download.

    JSON format:
    {
        "filename": "GanttChart_22052026",   // optional
        "dates": ["30/04", "01/05", ...],    // danh sách ngày header
        "rows": [
            {
                "parent": "Check update web",
                "task": "Check update web",
                "employee": "Nguyễn Thị Thanh Thanh",
                "status": 1,               // 1=done, 0=todo, 2=inprogress
                "total": 8.0,
                "hc": 8.0,
                "ot": 0.0,
                "date_hours": [8.0, 0, 0, ...]  // giờ theo từng ngày, cùng thứ tự dates
            }
        ],
        "grand_total": 120.0,
        "grand_hc": 100.0,
        "grand_ot": 20.0,
        "grand_date_hours": [8.0, 0, ...]   // tổng giờ từng ngày
    }
    """
    body = await request.json()

    dates           = body.get("dates", [])
    rows            = body.get("rows", [])
    grand_total     = body.get("grand_total", 0)
    grand_hc        = body.get("grand_hc", 0)
    grand_ot        = body.get("grand_ot", 0)
    grand_date_hours= body.get("grand_date_hours", [])
    filename        = body.get("filename", "GanttChart") + ".xlsx"

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Gantt Chart"

    # ── Cột cố định ──────────────────────────────────────────
    FIXED_COLS = ["Parent Task", "Task", "Nhân sự", "Trạng thái", "Tổng", "HC", "OT"]
    fixed_widths = [22, 28, 22, 18, 9, 9, 9]
    header_row = FIXED_COLS + dates

    # ── Viết header ──────────────────────────────────────────
    for col_idx, col_name in enumerate(header_row, start=1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.fill      = make_fill(COLOR["header_bg"])
        cell.font      = make_font(COLOR["header_fg"], bold=True)
        cell.alignment = make_align("center" if col_idx > 4 else "left")
        cell.border    = make_border()

    # Row height header
    ws.row_dimensions[1].height = 22

    # ── Viết data rows ────────────────────────────────────────
    STATUS_TEXT = {1: "HOÀN THÀNH", 0: "CHƯA THỰC HIỆN", 2: "ĐANG THỰC HIỆN"}

    for row_idx, row in enumerate(rows, start=2):
        status     = row.get("status", 2)
        total      = row.get("total", 0)
        hc         = row.get("hc", 0)
        ot         = row.get("ot", 0)
        date_hours = row.get("date_hours", [])

        # Màu trạng thái
        if status == 1:
            st_bg, st_fg = COLOR["done_bg"], COLOR["done_fg"]
        elif status == 0:
            st_bg, st_fg = COLOR["todo_bg"], COLOR["todo_fg"]
        else:
            st_bg, st_fg = COLOR["inprog_bg"], COLOR["inprog_fg"]

        row_data = [
            (row.get("parent", ""),   "FFFFFF", "000000", "left",   False),
            (row.get("task", ""),     "FFFFFF", "000000", "left",   False),
            (row.get("employee", ""), "FFFFFF", "000000", "left",   False),
            (STATUS_TEXT.get(status, ""), st_bg, st_fg, "center", True),
            (f"{total:.2f}h",         COLOR["total_bg"], COLOR["total_fg"], "center", True),
            (f"{hc:.2f}h",            COLOR["hc_bg"],    COLOR["hc_fg"],    "center", True),
            (f"{ot:.2f}h",            COLOR["ot_bg"],    COLOR["ot_fg"],    "center", True),
        ]

        # Cột cố định
        for col_idx, (val, bg, fg, align, bold) in enumerate(row_data, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.fill      = make_fill(bg)
            cell.font      = make_font(fg, bold=bold)
            cell.alignment = make_align(align)
            cell.border    = make_border()

        # Cột ngày
        for d_idx, hours in enumerate(date_hours):
            col_idx = 8 + d_idx
            if hours and hours > 0:
                val = f"{hours:.2f}h"
                bg  = COLOR["date_bg"]
            else:
                val = ""
                bg  = "FFFFFF"
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.fill      = make_fill(bg)
            cell.font      = make_font("000000")
            cell.alignment = make_align("center")
            cell.border    = make_border()

        ws.row_dimensions[row_idx].height = 20

    # ── Viết footer (TỔNG CỘNG) ───────────────────────────────
    footer_row = len(rows) + 2

    # Merge 4 cột đầu
    ws.merge_cells(
        start_row=footer_row, start_column=1,
        end_row=footer_row,   end_column=4
    )
    footer_cell = ws.cell(row=footer_row, column=1, value="TỔNG CỘNG")
    footer_cell.fill      = make_fill(COLOR["footer_bg"])
    footer_cell.font      = make_font(COLOR["footer_fg"], bold=True)
    footer_cell.alignment = make_align("left")
    footer_cell.border    = make_border()

    # Tổng / HC / OT footer
    footer_summary = [
        (f"{grand_total:.2f}h", COLOR["total_bg"], COLOR["total_fg"]),
        (f"{grand_hc:.2f}h",    COLOR["hc_bg"],    COLOR["hc_fg"]),
        (f"{grand_ot:.2f}h",    COLOR["ot_bg"],    COLOR["ot_fg"]),
    ]
    for i, (val, bg, fg) in enumerate(footer_summary):
        col_idx = 5 + i
        cell = ws.cell(row=footer_row, column=col_idx, value=val)
        cell.fill      = make_fill(bg)
        cell.font      = make_font(fg, bold=True)
        cell.alignment = make_align("center")
        cell.border    = make_border()

    # Cột ngày footer
    for d_idx, hours in enumerate(grand_date_hours):
        col_idx = 8 + d_idx
        if hours and hours > 0:
            val = f"{hours:.2f}h"
            bg  = COLOR["footer_date_bg"]
            fg  = "FFFFFF"
        else:
            val = ""
            bg  = COLOR["footer_bg"]
            fg  = "FFFFFF"
        cell = ws.cell(row=footer_row, column=col_idx, value=val)
        cell.fill      = make_fill(bg)
        cell.font      = make_font(fg, bold=True)
        cell.alignment = make_align("center")
        cell.border    = make_border()

    ws.row_dimensions[footer_row].height = 22

    # ── Column widths ─────────────────────────────────────────
    for i, w in enumerate(fixed_widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w
    for d_idx in range(len(dates)):
        ws.column_dimensions[get_column_letter(8 + d_idx)].width = 10

    # ── Freeze panes (7 cột đầu + 1 hàng header) ─────────────
    ws.freeze_panes = "H2"

    # ── Xuất file ─────────────────────────────────────────────
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )


@app.get("/health")
def health():
    return {"status": "ok"}
